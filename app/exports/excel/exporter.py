"""Deterministic XLSX export from analysis payload."""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook


def export_analysis_to_xlsx(analysis: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis"

    row = 1
    for section in [
        "inputs",
        "assumptions",
        "expert_load_contributions",
        "confidence",
        "profiles",
        "sizing",
        "bom",
        "performance",
        "economics",
        "co2",
        "warnings",
        "upgrade_paths",
    ]:
        ws.cell(row=row, column=1, value=section)
        ws.cell(row=row, column=2, value=str(analysis.get(section)))
        row += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
